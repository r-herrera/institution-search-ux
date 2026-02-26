#!/usr/bin/env python3
"""
Alternative export script: reads the source Excel file directly
and outputs a slim JSON file for the static search UI.

Usage:
    python3 scripts/export-from-excel.py --input ../shared/data/institutions.xlsx
    python3 scripts/export-from-excel.py --input ../shared/data/institutions.xlsx --output data/institutions.json

Requirements:
    pip install openpyxl
"""

import argparse
import json
import os
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def main():
    parser = argparse.ArgumentParser(description='Export institutions from Excel to JSON')
    parser.add_argument('--input', '-i', required=True, help='Path to institutions Excel file')
    parser.add_argument('--output', '-o', default='data/institutions.json', help='Output JSON file path')
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"Error: Input file not found: {args.input}")
        sys.exit(1)

    print(f"Loading workbook: {args.input}")
    wb = load_workbook(args.input, read_only=True)
    ws = wb.active

    # Read header row
    rows = ws.iter_rows(values_only=True)
    raw_headers = list(next(rows))

    # Build column index map: cleaned name -> first column index
    # (keeps duplicates by only recording the first occurrence per name)
    col_index: dict = {}
    for i, h in enumerate(raw_headers):
        if h is None:
            continue
        clean = (
            str(h).strip().lower()
            .replace(' ', '_').replace('/', '_')
            .replace('(', '').replace(')', '')
        )
        if clean not in col_index:
            col_index[clean] = i

    print(f"Column map (first 10): {list(col_index.items())[:10]}")

    def col_val(row: tuple, *names: str) -> 'str | None':
        """Return the first non-empty value for the given column name(s)."""
        for name in names:
            idx = col_index.get(name)
            if idx is not None and idx < len(row) and row[idx] is not None:
                v = str(row[idx]).strip()
                if v:
                    return v
        return None

    institutions = []
    record_id = 1

    for row_values in rows:
        row = row_values

        name = col_val(row, 'organisation_name')
        if not name:
            continue

        inst = {
            'id': record_id,
            'organisation_name': name,
            'city': col_val(row, 'city', 'city_town'),
            'country': col_val(row, 'country_asf', 'country'),
            'street_1': col_val(row, 'street_1'),
            'street_2': col_val(row, 'street_2'),
            'website': col_val(row, 'website'),
        }

        institutions.append(inst)
        record_id += 1

    wb.close()

    # Write output
    os.makedirs(os.path.dirname(args.output) or '.', exist_ok=True)
    with open(args.output, 'w', encoding='utf-8') as f:
        json.dump(institutions, f, ensure_ascii=False)

    size_mb = os.path.getsize(args.output) / (1024 * 1024)
    print(f"\nExport complete!")
    print(f"  Records: {len(institutions):,}")
    print(f"  File:    {args.output}")
    print(f"  Size:    {size_mb:.1f} MB")

    # Summary by country
    from collections import Counter
    countries = Counter(i['country'] for i in institutions)
    print(f'  Distinct countries: {len(countries)}')
    for c, n in countries.most_common(5):
        print(f'    {c}: {n:,}')


if __name__ == '__main__':
    main()
